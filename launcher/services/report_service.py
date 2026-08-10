import os
from datetime import datetime, date

from openpyxl import load_workbook


# ==========================================================
# ROOT PROJECT
# ==========================================================

PROJECT_ROOT = os.path.dirname(
    os.path.dirname(
        os.path.dirname(
            os.path.abspath(__file__)
        )
    )
)


# ==========================================================
# DATABASE
# ==========================================================

FILE = os.path.join(
    PROJECT_ROOT,
    "backend",
    "database",
    "TERATAI_CORE.xlsx"
)


class ReportService:

    # ======================================================
    # LOAD WORKBOOK
    # ======================================================

    def load_workbook(self):

        if not os.path.exists(FILE):

            raise FileNotFoundError(
                f"TERATAI_CORE.xlsx tidak ditemukan:\n{FILE}"
            )

        return load_workbook(
            FILE,
            data_only=True
        )


    # ======================================================
    # NORMALIZE DATE
    # ======================================================

    def normalize_date(self, value):

        if value is None:

            return None


        if isinstance(value, datetime):

            return value.date()


        if isinstance(value, date):

            return value


        if isinstance(value, str):

            value = value.strip()


            formats = [

                "%Y-%m-%d",

                "%Y-%m-%d %H:%M:%S",

                "%d/%m/%Y",

                "%d-%m-%Y"

            ]


            for date_format in formats:

                try:

                    return datetime.strptime(

                        value,

                        date_format

                    ).date()

                except ValueError:

                    continue


        return None


    # ======================================================
    # READ SHEET
    # ======================================================

    def read_sheet(

        self,

        workbook,

        sheet_name

    ):

        if sheet_name not in workbook.sheetnames:

            return []


        sheet = workbook[sheet_name]


        rows = list(

            sheet.iter_rows(

                values_only=True

            )

        )


        if not rows:

            return []


        headers = [

            str(value).strip()

            if value is not None

            else ""

            for value in rows[0]

        ]


        records = []


        for row in rows[1:]:

            record = {}


            for index, header in enumerate(headers):

                if not header:

                    continue


                value = (

                    row[index]

                    if index < len(row)

                    else None

                )


                record[header] = value


            records.append(record)


        return records


    # ======================================================
    # CHECK PERIOD
    # ======================================================

    def in_period(

        self,

        value,

        tanggal_mulai,

        tanggal_akhir

    ):

        tanggal = self.normalize_date(value)


        if not tanggal:

            return False


        return (

            tanggal_mulai

            <=

            tanggal

            <=

            tanggal_akhir

        )


    # ======================================================
    # GET SUMMARY
    # ======================================================

    def get_summary(

        self,

        tanggal_mulai,

        tanggal_akhir

    ):

        workbook = self.load_workbook()


        aktivitas = self.read_sheet(

            workbook,

            "AKTIVITAS"

        )


        aduan = self.read_sheet(

            workbook,

            "ADUAN"

        )


        # ==================================================
        # DATA AKTIVITAS
        # ==================================================

        total_aktivitas = 0


        # Semua pengguna unik

        pengguna_unik = set()


        # Pengguna administrasi unik

        pengguna_administrasi = set()


        # Total sesi administrasi

        sesi_administrasi = 0


        # Ringkasan jenis aktivitas

        aktivitas_per_jenis = {}


        for item in aktivitas:


            # ----------------------------------------------
            # FILTER PERIODE
            # ----------------------------------------------

            if not self.in_period(

                item.get("Waktu"),

                tanggal_mulai,

                tanggal_akhir

            ):

                continue


            total_aktivitas += 1


            # ----------------------------------------------
            # USER KEY
            # ----------------------------------------------

            user_key = (

                item.get("User_Key")

                or

                item.get("No_WA")

                or

                ""

            )


            if user_key:

                pengguna_unik.add(

                    str(user_key)

                )


            # ----------------------------------------------
            # JENIS AKTIVITAS
            # ----------------------------------------------

            jenis = str(

                item.get(

                    "Jenis_Aktivitas",

                    ""

                )

            ).strip().upper()


            if jenis:


                aktivitas_per_jenis[jenis] = (

                    aktivitas_per_jenis.get(

                        jenis,

                        0

                    )

                    + 1

                )


            # ----------------------------------------------
            # ADMINISTRASI
            # ----------------------------------------------

            if jenis == "ADMINISTRASI":


                sesi_administrasi += 1


                if user_key:


                    pengguna_administrasi.add(

                        str(user_key)

                    )


        # ==================================================
        # DATA ADUAN
        # ==================================================

        total_aduan = 0

        baru = 0

        diproses = 0

        menunggu_info = 0

        selesai = 0


        pengguna_aduan = set()


        for item in aduan:


            # ----------------------------------------------
            # FILTER PERIODE
            # ----------------------------------------------

            if not self.in_period(

                item.get("Tanggal"),

                tanggal_mulai,

                tanggal_akhir

            ):

                continue


            total_aduan += 1


            # ----------------------------------------------
            # USER ADUAN
            # ----------------------------------------------

            user_key = (

                item.get("User_Key")

                or

                item.get("No_WA")

                or

                ""

            )


            if user_key:

                pengguna_aduan.add(

                    str(user_key)

                )


            # ----------------------------------------------
            # STATUS
            # ----------------------------------------------

            status = str(

                item.get(

                    "Status",

                    ""

                )

            ).strip().upper()


            if status == "BARU":

                baru += 1


            elif status == "DIPROSES":

                diproses += 1


            elif status == "MENUNGGU_INFO":

                menunggu_info += 1


            elif status == "SELESAI":

                selesai += 1

        # ----------------------------------------------
        # REKAP PETUGAS
        # ----------------------------------------------

        aduan_per_petugas = {}

        for item in aduan:

            if not self.in_period(

                item.get("Tanggal"),

                tanggal_mulai,

                tanggal_akhir

            ):

                continue

            owner = str(

                    item.get(

                            "Owner"

                    )

                    or

                    item.get(

                            "Petugas"

                    )

                    or

                    "-"

            ).strip()


            if not owner:

                    owner = "Belum Ditugaskan"


            aduan_per_petugas[owner] = (

                    aduan_per_petugas.get(

                            owner,

                            0

                    )

                    + 1

            )

        # ==================================================
        # HASIL LAPORAN
        # ==================================================

        return {

            "periode": {

                "mulai": str(

                    tanggal_mulai

                ),

                "akhir": str(

                    tanggal_akhir

                )

            },


            "aktivitas": {

                "totalAktivitas":

                    total_aktivitas,


                "totalPenggunaUnik":

                    len(

                        pengguna_unik

                    ),


                "totalPenggunaAdministrasi":

                    len(

                        pengguna_administrasi

                    ),


                "totalSesiAdministrasi":

                    sesi_administrasi,


                "aktivitasPerJenis":

                    aktivitas_per_jenis

            },


            "aduan": {

                "totalAduan":

                    total_aduan,


                "totalPenggunaAduan":

                    len(

                        pengguna_aduan

                    ),


                "baru":

                    baru,


                "diproses":

                    diproses,


                "menungguInfo":

                    menunggu_info,


                "selesai":

                    selesai,

                "aduanPerPetugas":

                    aduan_per_petugas

            }

        }


# ==========================================================
# INSTANCE
# ==========================================================

report_service = ReportService()