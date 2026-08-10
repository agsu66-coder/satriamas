import os
from openpyxl import load_workbook


class ExcelService:

    def __init__(self):

        self.file_path = os.path.join(
            os.path.dirname(__file__),
            "..",
            "database",
            "TERATAI_CORE.xlsx"
        )

        self.workbook = None

    # ======================================================
    # Workbook
    # ======================================================

    def load_workbook(self):

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(
                f"Workbook tidak ditemukan : {self.file_path}"
            )

        self.workbook = load_workbook(self.file_path)

        return True

    def save_workbook(self):

        if self.workbook:
            self.workbook.save(self.file_path)

    def reload_workbook(self):

        self.load_workbook()

    def close_workbook(self):

        self.workbook = None

    # ======================================================
    # Sheet
    # ======================================================

    def get_sheet(self, sheet_name):

        if self.workbook is None:
            self.load_workbook()

        if sheet_name not in self.workbook.sheetnames:
            raise Exception(
                f"Sheet '{sheet_name}' tidak ditemukan."
            )

        return self.workbook[sheet_name]

    # ======================================================
    # Read
    # ======================================================

    def read_rows(self, sheet_name):

        sheet = self.get_sheet(sheet_name)

        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        return data

    def read_as_dict(self, sheet_name):
        """
        Membaca sheet berdasarkan header.

        Return:
        [
            {
                "KEY": "...",
                "Kategori": "...",
                "Isi_Pesan": "...",
                "Keterangan": "..."
            }
        ]

        Aman apabila:
        - urutan kolom berubah
        - ada kolom tambahan
        """

        rows = self.read_rows(sheet_name)

        if len(rows) < 2:
            return []

        headers = []

        for header in rows[0]:

            if header is None:
                headers.append("")
            else:
                headers.append(str(header).strip())

        data = []

        for row in rows[1:]:

            item = {}

            for i, header in enumerate(headers):

                if header == "":
                    continue

                value = row[i] if i < len(row) else None

                item[header] = value

            data.append(item)

        return data

    # ======================================================
    # Write
    # ======================================================

    def append_row(self, sheet_name, row_data):

        sheet = self.get_sheet(sheet_name)

        sheet.append(row_data)

        self.save_workbook()

    def update_cell(self, sheet_name, row, column, value):

        sheet = self.get_sheet(sheet_name)

        sheet.cell(
            row=row,
            column=column
        ).value = value

        self.save_workbook()

    # ======================================================
    # Utility
    # ======================================================

    def sheet_exists(self, sheet_name):

        if self.workbook is None:
            self.load_workbook()

        return sheet_name in self.workbook.sheetnames

    def get_sheet_names(self):

        if self.workbook is None:
            self.load_workbook()

        return self.workbook.sheetnames


excel_service = ExcelService()